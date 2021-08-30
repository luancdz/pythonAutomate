#import win32com.client
from screenShotClass import ScreenShot
import openpyxl
import time
import loginSAP
from runScripstClass import RunScripts
import constant
from tkinter import *
from tkinter import messagebox

def SAP_OP():

    #call class open SAP GUI
    sapGui  =  loginSAP.SapGui()

    pathExel = constant.FILENAME
    wb_obj = openpyxl.load_workbook(pathExel) 
    sheet_obj = wb_obj.active

    #total rows at spreadsheet
    numRow = sheet_obj.max_row

    for index in range( constant.INDEX_WITH_HEADER ,numRow + 1):

        try: 

            ovOrig = sheet_obj.cell(row = index, column = 1) 
            
            runScript = RunScripts(sapGui)

            #Get SO from PRD
            ovNew = runScript.getSOfromProd(ovOrig.value)

            #get from spreadsheet Type SO and set SO Ref
            tpOV = sheet_obj.cell(row = index, column = 4) 
            ovRef = sheet_obj.cell(row = index, column = 12)

            ovRef.value = runScript.createSOByRef(tpOV.value, ovNew)

            #shipPoint = sheet_obj.cell(row = 2, column = 7)
            wb_obj.save(pathExel)

            shipPoint = sheet_obj.cell(row = index, column = 7)

            #Define serial number/ set stock
            plant = sheet_obj.cell(row = index, column = 3)
            nrItem = sheet_obj.cell(row = index, column = 6)
            matnr = sheet_obj.cell(row = index, column = 2)
            qtdItem = sheet_obj.cell(row = index, column = 5)
            serNum = sheet_obj.cell(row = index, column = 8)
            stLocation = sheet_obj.cell(row = index, column = 11)
            stLocation.value = runScript.getStoredLoc(shipPoint.value, ovRef.value, index - 2)

            serNum.value = runScript.setStockAndSerNumber(ovRef.value, plant.value, nrItem.value, matnr.value, qtdItem.value, stLocation.value)

            #SAVE SERIAL NUMBER
            wb_obj.save(pathExel)

            #Delivery save 
            runScript.saveDelivery(shipPoint.value, ovRef.value)
            deliveryValue = runScript.getDelivery()

            #Save Delivery into SpreadSheet
            dlvNum = sheet_obj.cell(row = index, column = 13)
            dlvNum2 = sheet_obj.cell(row = index, column = 14)

            dlvNum.value = deliveryValue
            dlvNum2.value = deliveryValue
            wb_obj.save(pathExel)

            #Picking
            runScript.setDefPicking(qtdItem.value)

             #set serial number and post good issues
            runScript.setSerialNumAndPostGI(serNum.value)

            #get billing doc
            billDoc = sheet_obj.cell(row = index, column = 15)
            billDoc.value = runScript.createBilling()
            wb_obj.save(pathExel)

            #GET GOODS MVT DOC
            goodMov = sheet_obj.cell(row = index, column = 16)
            goodMov.value = runScript.getGoodsMvt(ovRef.value, nrItem.value, constant.GOOD_MOV)    
            wb_obj.save(pathExel)

        except:
            
            screenShot = ScreenShot()
            time.sleep(2)
            screenShot.screenShot(ovOrig.value)

            print('Error, verificar Screen Shots')

        finally:
            continue

 
    messagebox.showinfo("Show info", "Script Finalizado")
    sapGui.killSAP()

    # while sapGui.connection.children.count > 0:
    #     time.sleep(10)
    #     print("SAP GUI OPEN")


    # session = None
    # connection = None
    # application = None
    # SapGuiAuto = None

if __name__ == '__main__':
    window = Tk()
    window.geometry("200x75")
    window.title("NF Automate SAP")
    button = Button(window, text='AG3 Login', command= lambda : SAP_OP())
    button.pack()
    mainloop()

#SAP_OP()