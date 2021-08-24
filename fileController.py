from tkinter import filedialog
from tkinter import filedialog
from tkinter.constants import SOLID
import openpyxl
import numpy as np

def getPath():

    return filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx; *.xls")])


def getSO(path):

    def getSONumber(m_row):

        SOList = []

        # Loop will print all values
        # of first column
        for i in range(2, m_row + 1):
            cell_obj = sheet_obj.cell(row = i, column = 1)
            SOList.append(cell_obj.value)

        return SOList

    
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active

    m_row = sheet_obj.max_row

    return getSONumber(m_row)
 

